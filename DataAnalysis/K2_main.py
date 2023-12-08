# -*- coding: utf-8 -*-
import os
import sys
import tkinter
from tkinter import filedialog, ttk
import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QWidget
from openpyxl.styles import NamedStyle
from tqdm import tk

from DataAnalysis.test import chi_square_test
from Cross import Cross
from K2_result import K2_result
from show_table import Show_table


# Form implementation generated from reading ui file 'K2_main.ui'
#
# Created by: PyQt5 UI code generator 5.15.9
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(754, 530)
        # 加载图片
        image_path = "button.png"
        pixmap = QtGui.QPixmap(image_path)

        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.listWidget = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget.setGeometry(QtCore.QRect(20, 100, 211, 321))
        self.listWidget.setObjectName("listWidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(20, 80, 54, 12))
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(20, 20, 54, 31))
        self.label_2.setObjectName("label_2")
        self.lineEdit_file = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_file.setGeometry(QtCore.QRect(80, 20, 401, 31))
        self.lineEdit_file.setObjectName("lineEdit_file")
        self.pushButton_select_file = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_select_file.setGeometry(QtCore.QRect(500, 20, 91, 31))
        self.pushButton_select_file.setObjectName("pushButton_select_file")
        self.pushButton_select_file.clicked.connect(self.select_file)
        self.pushButton_1 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_1.setGeometry(QtCore.QRect(250, 120, 41, 41))
        self.pushButton_1.setObjectName("pushButton_1")
        self.pushButton_1.setIcon(QtGui.QIcon(pixmap))
        self.pushButton_1.setIconSize(QtCore.QSize(41, 41))
        self.pushButton_1.setStyleSheet(
            "QPushButton { border: none; background-image: url('button.png'); background-position: center; background-repeat: none; }")
        self.pushButton_1.clicked.connect(self.select_item1)
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(250, 240, 41, 41))
        self.pushButton_2.setIcon(QtGui.QIcon(pixmap))
        self.pushButton_2.setIconSize(QtCore.QSize(41, 41))
        self.pushButton_2.setStyleSheet(
            "QPushButton { border: none; background-image: url('button.png'); background-position: center; background-repeat: none; }")
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.clicked.connect(self.select_item2)
        self.pushButton_3 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_3.setGeometry(QtCore.QRect(250, 360, 41, 41))
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_3.setIcon(QtGui.QIcon(pixmap))
        self.pushButton_3.setIconSize(QtCore.QSize(41, 41))
        self.pushButton_3.setStyleSheet(
            "QPushButton { border: none; background-image: url('button.png'); background-position: center; background-repeat: none; }")
        self.pushButton_3.clicked.connect(self.select_item3)
        self.listWidget_show = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget_show.setGeometry(QtCore.QRect(310, 100, 161, 81))
        self.listWidget_show.setObjectName("listWidget_show")
        self.listWidget_show.itemDoubleClicked.connect(self.remove_item1)
        self.listWidget_row = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget_row.setGeometry(QtCore.QRect(310, 220, 161, 81))
        self.listWidget_row.setObjectName("listWidget_row")
        self.listWidget_row.itemDoubleClicked.connect(self.remove_item2)
        self.listWidget_col = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget_col.setGeometry(QtCore.QRect(310, 340, 161, 81))
        self.listWidget_col.setObjectName("listWidget_col")
        self.listWidget_col.itemDoubleClicked.connect(self.remove_item3)
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(310, 80, 161, 20))
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(310, 200, 161, 20))
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        self.label_5.setGeometry(QtCore.QRect(310, 320, 161, 20))
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self.centralwidget)
        self.label_6.setGeometry(QtCore.QRect(500, 80, 54, 20))
        self.label_6.setObjectName("label_6")
        self.recode_Button = QtWidgets.QPushButton(self.centralwidget)
        self.recode_Button.setGeometry(QtCore.QRect(650, 100, 81, 31))
        self.recode_Button.setObjectName("recode_Button")
        self.recode_Button.clicked.connect(self.recode)
        self.label_7 = QtWidgets.QLabel(self.centralwidget)
        self.label_7.setGeometry(QtCore.QRect(500, 140, 121, 31))
        self.label_7.setObjectName("label_7")
        self.radioButton = QtWidgets.QRadioButton(self.centralwidget)
        self.radioButton.setGeometry(QtCore.QRect(500, 180, 51, 21))
        self.radioButton.setObjectName("radioButton")
        self.radioButton_2 = QtWidgets.QRadioButton(self.centralwidget)
        self.radioButton_2.setGeometry(QtCore.QRect(500, 220, 51, 21))
        self.radioButton_2.setObjectName("radioButton_2")
        self.label_8 = QtWidgets.QLabel(self.centralwidget)
        self.label_8.setGeometry(QtCore.QRect(520, 250, 31, 31))
        self.label_8.setObjectName("label_8")
        self.pushButton_add = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_add.setGeometry(QtCore.QRect(650, 250, 81, 31))
        self.pushButton_add.setObjectName("pushButton_add")
        self.pushButton_add.clicked.connect(self.add)
        self.label_9 = QtWidgets.QLabel(self.centralwidget)
        self.label_9.setGeometry(QtCore.QRect(500, 290, 51, 31))
        self.label_9.setObjectName("label_9")
        self.lineEdit_new = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_new.setGeometry(QtCore.QRect(560, 250, 81, 31))
        self.lineEdit_new.setObjectName("lineEdit_new")
        self.lineEdit_low = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_low.setGeometry(QtCore.QRect(560, 210, 81, 31))
        self.lineEdit_low.setObjectName("lineEdit_low")
        self.lineEdit_old = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_old.setGeometry(QtCore.QRect(560, 170, 81, 31))
        self.lineEdit_old.setObjectName("lineEdit_old")
        self.lineEdit_up = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_up.setGeometry(QtCore.QRect(650, 210, 81, 31))
        self.lineEdit_up.setObjectName("lineEdit_up")
        self.listWidget_show_recode = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget_show_recode.setGeometry(QtCore.QRect(560, 290, 171, 131))
        self.listWidget_show_recode.setObjectName("listWidget_show_recode")
        self.submit_Button = QtWidgets.QPushButton(self.centralwidget)
        self.submit_Button.setGeometry(QtCore.QRect(70, 440, 75, 41))
        self.submit_Button.setObjectName("submit_Button")
        self.submit_Button.clicked.connect(self.Submit)
        self.paste_Button = QtWidgets.QPushButton(self.centralwidget)
        self.paste_Button.setGeometry(QtCore.QRect(200, 440, 75, 41))
        self.paste_Button.setObjectName("paste_Button")
        self.paste_Button.clicked.connect(self.Paste)
        self.reset_Button = QtWidgets.QPushButton(self.centralwidget)
        self.reset_Button.setGeometry(QtCore.QRect(330, 440, 81, 41))
        self.reset_Button.setObjectName("reset_Button")
        self.reset_Button.clicked.connect(self.Reset)
        self.cancel_Button = QtWidgets.QPushButton(self.centralwidget)
        self.cancel_Button.setGeometry(QtCore.QRect(470, 440, 81, 41))
        self.cancel_Button.setObjectName("cancel_Button")
        self.cancel_Button.clicked.connect(self.Cancel)
        self.help_Button_ = QtWidgets.QPushButton(self.centralwidget)
        self.help_Button_.setGeometry(QtCore.QRect(610, 440, 81, 41))
        self.help_Button_.setObjectName("help_Button_")
        self.help_Button_.clicked.connect(self.Help)
        self.lineEdit_recode = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_recode.setGeometry(QtCore.QRect(500, 100, 141, 31))
        self.lineEdit_recode.setObjectName("lineEdit_recode")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 754, 23))
        self.menubar.setObjectName("menubar")
        self.menuF = QtWidgets.QMenu(self.menubar)
        self.menuF.setObjectName("menuF")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.menubar.addAction(self.menuF.menuAction())
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.label.setText(_translate("MainWindow", "列项"))
        self.label_2.setText(_translate("MainWindow", "文件名称:"))
        self.pushButton_select_file.setText(_translate("MainWindow", "选择文件"))
        self.pushButton_1.setText(_translate("MainWindow", ""))
        self.pushButton_2.setText(_translate("MainWindow", ""))
        self.pushButton_3.setText(_translate("MainWindow", ""))
        self.label_3.setText(_translate("MainWindow", "数字变量-->输出变量:"))
        self.label_4.setText(_translate("MainWindow", "行(O):"))
        self.label_5.setText(_translate("MainWindow", "列(C):"))
        self.label_6.setText(_translate("MainWindow", "输出变量"))
        self.recode_Button.setText(_translate("MainWindow", "重新编码"))
        self.label_7.setText(_translate("MainWindow", "选择旧值并设置新值"))
        self.radioButton.setText(_translate("MainWindow", "值："))
        self.radioButton_2.setText(_translate("MainWindow", "范围："))
        self.label_8.setText(_translate("MainWindow", "新值："))
        self.pushButton_add.setText(_translate("MainWindow", "添加"))
        self.label_9.setText(_translate("MainWindow", "旧-->新："))
        self.submit_Button.setText(_translate("MainWindow", "确定"))
        self.paste_Button.setText(_translate("MainWindow", "粘贴"))
        self.reset_Button.setText(_translate("MainWindow", "重置"))
        self.cancel_Button.setText(_translate("MainWindow", "取消"))
        self.help_Button_.setText(_translate("MainWindow", "帮助"))
        self.menuF.setTitle(_translate("MainWindow", "F检验"))

    def select_file(self):
        try:
            file_dialog = QFileDialog()
            file_name, _ = file_dialog.getOpenFileName(self.centralwidget, "Select File",
                                                       "", "Excel Files (*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv)")
            global global_file_name
            global_file_name = file_name
            # 打开Excel文件
            workbook = openpyxl.load_workbook(file_name)
            # 将文件名设置为文本框的值
            file_name = os.path.basename(file_name)
            self.lineEdit_file.clear()  # 清空文本框内容
            self.lineEdit_file.insert(file_name)  # 设置文本框的值
            # 添加默认样式
            default_style = NamedStyle(name="default")
            workbook.default_style = default_style
            # 获取第一个工作表
            sheet = workbook.active
            global glo_sheet
            glo_sheet = sheet
            # 获取列名
            column_names = [cell.value for cell in sheet[1]]
            global selected_column_names
            selected_column_names = column_names
            # 清空listbox内容
            self.listWidget.clear()
            # 将列名添加到listbox
            for name in column_names:
                self.listWidget.addItem(name)

        except Exception as e:
            # 处理异常
            print("选择文件时出现错误:", str(e))

    def Submit(self):
        row_items = [self.listWidget_row.item(i).text() for i in range(self.listWidget_row.count())]
        col_items = [self.listWidget_col.item(i).text() for i in range(self.listWidget_row.count())]
        # selected_items = [item.strip() for item in selected_items]  # 去除换行符
        print(f'row_items:{row_items},col_items:{col_items}')
        # 打开Excel文件
        workbook = openpyxl.load_workbook(global_file_name)
        # 获取第一个工作表
        sheet = workbook.active
        # 获取选中列的全部值并存储在二维数组中
        row_values = [[] for _ in range(len(row_items))]
        col_values = [[] for _ in range(len(col_items))]
        # 获取行和列值
        for i, row_name in enumerate(row_items):
            for j, col_name in enumerate(col_items):
                selected_row_name = row_name
                row_index = selected_column_names.index(selected_row_name) + 1  # 注意索引从1开始
                row_values_ = [cell[0] for cell in sheet.iter_rows(min_row=2, min_col=row_index, values_only=True)]
                row_values[i] = row_values_
                print(f'选择的行名{selected_row_name}\n行所在索引{row_index}\n对应excel表中值{row_values_}\n')

                if col_name in new_col:
                    col_values[j] = new_col_value
                    print(f'获取到的列值(分组过):{col_values[j]}')
                else:
                    selected_column_name = col_name
                    column_index = selected_column_names.index(selected_column_name) + 1  # 注意索引从1开始
                    column_values = [cell[0] for cell in
                                     sheet.iter_rows(min_row=2, min_col=column_index, values_only=True)]
                    col_values[j] = column_values
                    print(
                        f'选择的列名{selected_column_name}\n列所在索引{column_index}\n对应excel表中值{column_values}\n')
                # 计算交叉表
                self.cross_window = Cross()
                self.cross_window.setupUi(row_name, col_name, row_values[i], col_values[j])
                self.cross_window.show()
                # 卡方检验结果
                self.result_window = K2_result()
                self.result_window.setupUi(row_values[i], col_values[j])
                self.result_window.show()
        # 展示表格内容
        self.show_table = Show_table()
        self.show_table.setupUi(global_file_name, glo_sheet, new_col=new_col, new_col_value=new_col_value)
        self.show_table.show()

    def select_item1(self):
        selected_items = self.listWidget.selectedItems()
        for item in selected_items:
            global old_item
            old_item = item
            old_column_name = old_item.text()
            print(f'列名:{old_column_name}')  # 肺活量
            # 获取旧列的索引
            column_index = selected_column_names.index(old_column_name) + 1  # 注意索引从1开始
            print(f'excel中对应索引:{column_index}')  # 7
            # 获取旧列的全部值
            column_values = [cell[0] for cell in
                             glo_sheet.iter_rows(min_row=2, min_col=column_index, values_only=True)]
            global new_col_value
            new_col_value = column_values
            # item.setText(item.text() + " --> ?")
            self.listWidget_show.addItem(item.text() + " --> ?")

    def remove_item1(self, item):
        row_index = self.listWidget_show.row(item)
        self.listWidget_show.takeItem(row_index)

    def select_item2(self):
        selected_items = self.listWidget.selectedItems()
        for item in selected_items:
            self.listWidget.takeItem(self.listWidget.row(item))
            self.listWidget_row.addItem(item.text())

    def remove_item2(self, item):
        row_index = self.listWidget_row.row(item)
        self.listWidget_row.takeItem(row_index)
        self.listWidget.addItem(item.text())

    def select_item3(self):
        selected_items = self.listWidget.selectedItems()
        for item in selected_items:
            self.listWidget.takeItem(self.listWidget.row(item))
            self.listWidget_col.addItem(item.text())

    def remove_item3(self, item):
        row_index = self.listWidget_col.row(item)
        self.listWidget_col.takeItem(row_index)
        self.listWidget.addItem(item.text())

    def recode(self):
        search_text = " --> ?"
        replace_text = " --> " + self.lineEdit_recode.text()
        global new_col
        new_col = self.lineEdit_recode.text()
        self.listWidget.addItem(self.lineEdit_recode.text())
        for index in range(self.listWidget_show.count()):
            item = self.listWidget_show.item(index)
            item_text = item.text()
            if item_text.endswith(search_text):
                new_text = item_text[:-len(search_text)] + replace_text
                item.setText(new_text)
                break
        self.lineEdit_recode.clear()

    def add(self):
        new_value = self.lineEdit_new.text()
        # 旧值替换新值
        if self.radioButton.isChecked():  # radioButton is selected
            old_value = self.lineEdit_old.text()
            for i, value in enumerate(new_col_value):
                if int(old_value) == int(new_col_value[i]):
                    new_col_value[i] = new_value
            print(f'新值:{new_col_value}')
            self.listWidget_show_recode.addItem(old_value + ' --> ' + new_value)
        # 取旧值范围为新值
        elif self.radioButton_2.isChecked():  # radioButton_2 is selected
            low_value = self.lineEdit_low.text()
            up_value = self.lineEdit_up.text()

            for i, value in enumerate(new_col_value):
                if int(low_value) <= int(new_col_value[i]) < int(up_value):
                    new_col_value[i] = new_value
            print(f'新值:{new_col_value}')
            self.listWidget_show_recode.addItem(low_value + ' thu ' + up_value + ' --> ' + new_value)
        self.lineEdit_old.clear()
        self.lineEdit_low.clear()
        self.lineEdit_up.clear()
        self.lineEdit_new.clear()

    def Help(self):
        print("帮助")

    def Cancel(self):
        print("取消")

    def Paste(self):
        print("粘贴")

    def Reset(self):
        print("重置")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    MainWindow = QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
